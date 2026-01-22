"""
Test Capture Tool - Records user interactions and generates test cases
Supports Chrome, Firefox, and Edge browsers with Normal/Private modes
"""

import json
import os
from datetime import datetime
from typing import List, Dict, Any
from playwright.sync_api import sync_playwright, Page, Browser, BrowserContext
import time
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class TestCaptureTool:
    def __init__(self):
        self.recorded_actions: List[Dict[str, Any]] = []
        self.browser: Browser = None
        self.context: BrowserContext = None
        self.page: Page = None
        self.playwright = None
        self.current_url = ""
        
    def get_browser_choice(self) -> tuple[str, str]:
        """Get browser and mode choice from user"""
        print("\n" + "="*60)
        print("Test Capture Tool - Browser Selection")
        print("="*60)
        print("\nSelect Browser:")
        print("1. Chrome")
        print("2. Firefox")
        print("3. Edge")
        
        browser_choice = input("\nEnter your choice (1/2/3): ").strip()
        
        browser_map = {
            "1": ("chrome", "Chrome"),
            "2": ("firefox", "Firefox"),
            "3": ("msedge", "Edge")
        }
        
        if browser_choice not in browser_map:
            print("Invalid choice. Defaulting to Chrome.")
            browser_choice = "1"
        
        browser_type, browser_name = browser_map[browser_choice]
        
        # Get mode choice
        print(f"\n{browser_name} Mode Selection:")
        if browser_name == "Chrome":
            print("1. Normal Mode")
            print("2. Incognito Mode")
            mode_choice = input("\nEnter your choice (1/2): ").strip()
            mode = "incognito" if mode_choice == "2" else "normal"
        else:  # Firefox or Edge
            print("1. Normal Mode")
            print("2. Private Mode")
            mode_choice = input("\nEnter your choice (1/2): ").strip()
            mode = "private" if mode_choice == "2" else "normal"
        
        return browser_type, mode
    
    def open_browser(self, browser_type: str, mode: str):
        """Open browser with specified type and mode"""
        print(f"\nOpening browser...")
        self.playwright = sync_playwright().start()
        
        # Launch browser based on type
        if browser_type == "chrome":
            browser_launcher = self.playwright.chromium
            # Use Chrome channel to launch actual Google Chrome browser
            launch_options = {
                "headless": False,
                "slow_mo": 100,
                "channel": "chrome"  # Use actual Google Chrome
            }
        elif browser_type == "firefox":
            browser_launcher = self.playwright.firefox
            launch_options = {
                "headless": False,
                "slow_mo": 100
            }
        elif browser_type == "msedge":
            # Edge uses chromium engine
            browser_launcher = self.playwright.chromium
            launch_options = {
                "headless": False,
                "slow_mo": 100,
                "channel": "msedge"
            }
        else:
            browser_launcher = self.playwright.chromium
            launch_options = {
                "headless": False,
                "slow_mo": 100
            }
        
        # Add private mode arguments
        if mode in ["incognito", "private"]:
            if browser_type == "chrome":
                if "args" not in launch_options:
                    launch_options["args"] = []
                launch_options["args"].append("--incognito")
            elif browser_type == "msedge":
                if "args" not in launch_options:
                    launch_options["args"] = []
                launch_options["args"].append("--inprivate")
            # Firefox private mode is handled via context
        
        self.browser = browser_launcher.launch(**launch_options)
        
        # Create context with or without private mode
        context_options = {
            "viewport": {"width": 1920, "height": 1080},
            "record_video_dir": "test_recordings/"
        }
        
        # For Firefox private mode, use private browsing context
        if mode == "private" and browser_type == "firefox":
            context_options["is_mobile"] = False
            # Firefox private browsing is enabled by default in isolated contexts
        
        # Create context (all Playwright contexts are isolated, simulating private mode)
        self.context = self.browser.new_context(**context_options)
        self.page = self.context.new_page()
        
        print(f"Browser opened successfully in {mode} mode!")
    
    def get_url(self) -> str:
        """Get URL from user"""
        print("\n" + "="*60)
        url = input("Enter the URL to test: ").strip()
        
        # Add protocol if missing
        if not url.startswith(("http://", "https://")):
            url = "https://" + url
        
        return url
    
    def navigate_to_url(self, url: str):
        """Navigate to the specified URL"""
        print(f"\nNavigating to: {url}")
        self.current_url = url
        self.page.goto(url, wait_until="networkidle")
        
        # Record initial page load
        self.record_action({
            "action": "navigate",
            "url": url,
            "timestamp": datetime.now().isoformat(),
            "page_title": self.page.title(),
            "page_url": self.page.url
        })
        
        print(f"Page loaded: {self.page.title()}")
        print("\n" + "="*60)
        print("Test Recording Started!")
        print("="*60)
        print("\nInstructions:")
        print("- Interact with the page (click, type, navigate)")
        print("- Type 'capture' to capture current page state")
        print("- Type 'stop' to stop recording and generate test cases")
        print("- Type 'help' for more commands")
        print("="*60 + "\n")
    
    def capture_page_state(self):
        """Capture current page state (elements, text, buttons)"""
        try:
            # Get page information
            page_info = {
                "url": self.page.url,
                "title": self.page.title(),
                "timestamp": datetime.now().isoformat()
            }
            
            # Capture all buttons
            buttons = self.page.query_selector_all("button, input[type='button'], input[type='submit'], a[role='button']")
            button_info = []
            for btn in buttons[:50]:  # Limit to 50 buttons
                try:
                    btn_text = btn.inner_text() or btn.get_attribute("value") or btn.get_attribute("aria-label") or ""
                    btn_id = btn.get_attribute("id") or ""
                    btn_class = btn.get_attribute("class") or ""
                    button_info.append({
                        "text": btn_text.strip(),
                        "id": btn_id,
                        "class": btn_class,
                        "tag": btn.evaluate("el => el.tagName.toLowerCase()")
                    })
                except:
                    continue
            
            # Capture all input fields
            inputs = self.page.query_selector_all("input, textarea, select")
            input_info = []
            for inp in inputs[:50]:  # Limit to 50 inputs
                try:
                    inp_type = inp.get_attribute("type") or "text"
                    inp_id = inp.get_attribute("id") or ""
                    inp_name = inp.get_attribute("name") or ""
                    inp_placeholder = inp.get_attribute("placeholder") or ""
                    input_info.append({
                        "type": inp_type,
                        "id": inp_id,
                        "name": inp_name,
                        "placeholder": inp_placeholder
                    })
                except:
                    continue
            
            # Capture visible text (first 1000 characters)
            try:
                body_text = self.page.inner_text("body")[:1000]
            except:
                body_text = ""
            
            page_state = {
                "action": "page_capture",
                **page_info,
                "buttons": button_info,
                "inputs": input_info,
                "body_text_preview": body_text
            }
            
            self.record_action(page_state)
            print(f"\n✓ Page state captured: {page_info['title']}")
            print(f"  - Found {len(button_info)} buttons")
            print(f"  - Found {len(input_info)} input fields")
            
        except Exception as e:
            print(f"Error capturing page state: {str(e)}")
    
    def record_action(self, action_data: Dict[str, Any]):
        """Record an action to the test case list"""
        self.recorded_actions.append(action_data)
    
    def setup_listeners(self):
        """Setup event listeners to automatically capture interactions"""
        last_url = self.page.url
        
        # Listen for navigation
        def handle_navigation(event):
            try:
                current_url = self.page.url
                if current_url != last_url:
                    self.record_action({
                        "action": "navigation",
                        "url": current_url,
                        "title": self.page.title(),
                        "timestamp": datetime.now().isoformat()
                    })
                    print(f"\n→ Navigation detected: {self.page.title()}")
            except:
                pass
        
        self.page.on("framenavigated", handle_navigation)
        
        # Note: For click and input events, we'll rely on manual capture
        # as automatic capture can be too verbose. Users can type 'capture' 
        # after performing actions to record the page state.
    
    def start_recording(self):
        """Start the interactive recording session"""
        self.setup_listeners()
        
        # Initial page capture
        self.capture_page_state()
        
        while True:
            try:
                command = input("\nEnter command (capture/stop/help): ").strip().lower()
                
                if command == "stop":
                    break
                elif command == "capture":
                    self.capture_page_state()
                elif command == "help":
                    print("\nAvailable Commands:")
                    print("  capture - Capture current page state (buttons, inputs, text)")
                    print("  stop    - Stop recording and generate test cases")
                    print("  help    - Show this help message")
                elif command == "":
                    # Allow user to interact with browser, then they can type 'capture'
                    print("(Interact with the browser, then type 'capture' to record)")
                else:
                    print("Unknown command. Type 'help' for available commands.")
                
                # Small delay to allow page interactions
                time.sleep(0.5)
                
            except KeyboardInterrupt:
                print("\n\nRecording interrupted by user.")
                break
            except Exception as e:
                print(f"Error during recording: {str(e)}")
                continue
    
    def generate_test_cases(self):
        """Generate test case file from recorded actions"""
        if not self.recorded_actions:
            print("No actions recorded. Nothing to generate.")
            return
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = "test_cases"
        os.makedirs(output_dir, exist_ok=True)
        
        # Generate JSON test case file
        json_file = os.path.join(output_dir, f"test_case_{timestamp}.json")
        with open(json_file, "w", encoding="utf-8") as f:
            json.dump({
                "test_case_id": f"TC_{timestamp}",
                "created_at": datetime.now().isoformat(),
                "initial_url": self.current_url,
                "total_actions": len(self.recorded_actions),
                "actions": self.recorded_actions
            }, f, indent=2, ensure_ascii=False)
        
        # Generate human-readable test case file
        txt_file = os.path.join(output_dir, f"test_case_{timestamp}.txt")
        with open(txt_file, "w", encoding="utf-8") as f:
            f.write("="*80 + "\n")
            f.write("TEST CASE DOCUMENTATION\n")
            f.write("="*80 + "\n\n")
            f.write(f"Test Case ID: TC_{timestamp}\n")
            f.write(f"Created At: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Initial URL: {self.current_url}\n")
            f.write(f"Total Actions Recorded: {len(self.recorded_actions)}\n")
            f.write("\n" + "="*80 + "\n")
            f.write("RECORDED ACTIONS\n")
            f.write("="*80 + "\n\n")
            
            for idx, action in enumerate(self.recorded_actions, 1):
                f.write(f"\n--- Action {idx} ---\n")
                f.write(f"Type: {action.get('action', 'unknown')}\n")
                f.write(f"Timestamp: {action.get('timestamp', 'N/A')}\n")
                
                if action.get('action') == 'navigate':
                    f.write(f"URL: {action.get('url')}\n")
                    f.write(f"Page Title: {action.get('page_title')}\n")
                
                elif action.get('action') == 'page_capture':
                    f.write(f"Page URL: {action.get('url')}\n")
                    f.write(f"Page Title: {action.get('title')}\n")
                    f.write(f"\nButtons Found ({len(action.get('buttons', []))}):\n")
                    for btn in action.get('buttons', [])[:10]:  # Show first 10
                        f.write(f"  - {btn.get('text', 'N/A')} (ID: {btn.get('id', 'N/A')})\n")
                    
                    f.write(f"\nInput Fields Found ({len(action.get('inputs', []))}):\n")
                    for inp in action.get('inputs', [])[:10]:  # Show first 10
                        f.write(f"  - Type: {inp.get('type')}, ID: {inp.get('id', 'N/A')}, Name: {inp.get('name', 'N/A')}\n")
                    
                    if action.get('body_text_preview'):
                        f.write(f"\nPage Text Preview:\n{action.get('body_text_preview')[:200]}...\n")
                
                elif action.get('action') == 'navigation':
                    f.write(f"Navigated to: {action.get('url')}\n")
                    f.write(f"New Page Title: {action.get('title')}\n")
                
                f.write("\n")
        
        # Generate Python test script (Playwright)
        py_file = os.path.join(output_dir, f"test_case_{timestamp}.py")
        with open(py_file, "w", encoding="utf-8") as f:
            f.write("""\"\"\"
Auto-generated Test Case
Generated by Test Capture Tool
\"\"\"

from playwright.sync_api import sync_playwright
import pytest


def test_recorded_actions():
    \"\"\"Test case generated from recorded user interactions\"\"\"
    with sync_playwright() as p:
        # Launch browser
        browser = p.chromium.launch(headless=False, channel="chrome")
        context = browser.new_context(viewport={"width": 1920, "height": 1080})
        page = context.new_page()
        
        try:
""")
            
            for action in self.recorded_actions:
                if action.get('action') == 'navigate':
                    f.write(f"            # Navigate to: {action.get('url')}\n")
                    f.write(f"            page.goto('{action.get('url')}', wait_until='networkidle')\n")
                    f.write(f"            assert page.title() == '{action.get('page_title', '')}'\n\n")
                
                elif action.get('action') == 'page_capture':
                    f.write(f"            # Page capture: {action.get('title')}\n")
                    f.write(f"            page.goto('{action.get('url')}', wait_until='networkidle')\n")
                    f.write(f"            # Verify page loaded\n")
                    f.write(f"            assert '{action.get('title', '')}' in page.title()\n\n")
                    
                    # Add assertions for buttons
                    for btn in action.get('buttons', [])[:5]:
                        if btn.get('id'):
                            f.write(f"            # Verify button: {btn.get('text', 'N/A')}\n")
                            f.write(f"            assert page.locator('#{btn.get('id')}').is_visible()\n\n")
            
            f.write("""        finally:
            context.close()
            browser.close()


if __name__ == "__main__":
    test_recorded_actions()
""")
        
        # Generate Excel test case file
        excel_file = None
        if EXCEL_AVAILABLE:
            excel_file = self.generate_excel_file(output_dir, timestamp)
        else:
            print("\n⚠ Warning: openpyxl not installed. Excel file will not be generated.")
            print("   Install it with: pip install openpyxl")
        
        print("\n" + "="*60)
        print("Test Cases Generated Successfully!")
        print("="*60)
        print(f"\nGenerated Files:")
        print(f"  1. JSON: {json_file}")
        print(f"  2. Text: {txt_file}")
        print(f"  3. Python: {py_file}")
        if excel_file:
            print(f"  4. Excel: {excel_file}")
        print(f"\nTotal Actions Recorded: {len(self.recorded_actions)}")
        print("="*60)
    
    def generate_excel_file(self, output_dir: str, timestamp: str) -> str:
        """Generate Excel file with test case data"""
        excel_file = os.path.join(output_dir, "Doceree_TestCases.xlsx")
        wb = Workbook()
        
        # Remove default sheet and create custom sheets
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Sheet 1: Test Case Summary
        ws_summary = wb.create_sheet("Test Case Summary", 0)
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        
        # Test Case Information
        ws_summary.merge_cells("A1:B1")
        ws_summary["A1"] = "Doceree Test Cases"
        ws_summary["A1"].font = title_font
        ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
        
        ws_summary["A3"] = "Test Case ID:"
        ws_summary["B3"] = f"TC_{timestamp}"
        ws_summary["A4"] = "Created At:"
        ws_summary["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_summary["A5"] = "Initial URL:"
        ws_summary["B5"] = self.current_url
        ws_summary["A6"] = "Total Actions:"
        ws_summary["B6"] = len(self.recorded_actions)
        
        # Sheet 2: Actions
        ws_actions = wb.create_sheet("Actions", 1)
        headers_actions = ["Action #", "Action Type", "Timestamp", "URL", "Page Title", "Description"]
        ws_actions.append(headers_actions)
        
        # Style headers
        for col_num, header in enumerate(headers_actions, 1):
            cell = ws_actions.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add action data
        for idx, action in enumerate(self.recorded_actions, 1):
            row = [
                idx,
                action.get('action', 'unknown'),
                action.get('timestamp', ''),
                action.get('url', ''),
                action.get('title') or action.get('page_title', ''),
                self._get_action_description(action)
            ]
            ws_actions.append(row)
        
        # Auto-adjust column widths
        for col in ws_actions.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_actions.column_dimensions[col_letter].width = adjusted_width
        
        # Sheet 3: Page Elements
        ws_elements = wb.create_sheet("Page Elements", 2)
        headers_elements = ["Action #", "Element Type", "Text/Value", "ID", "Class", "Name", "Type", "Page URL"]
        ws_elements.append(headers_elements)
        
        # Style headers
        for col_num, header in enumerate(headers_elements, 1):
            cell = ws_elements.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add element data
        action_num = 0
        for action in self.recorded_actions:
            if action.get('action') == 'page_capture':
                action_num += 1
                page_url = action.get('url', '')
                
                # Add buttons
                for btn in action.get('buttons', []):
                    row = [
                        action_num,
                        "Button",
                        btn.get('text', ''),
                        btn.get('id', ''),
                        btn.get('class', ''),
                        '',
                        btn.get('tag', ''),
                        page_url
                    ]
                    ws_elements.append(row)
                
                # Add inputs
                for inp in action.get('inputs', []):
                    row = [
                        action_num,
                        "Input",
                        inp.get('placeholder', ''),
                        inp.get('id', ''),
                        '',
                        inp.get('name', ''),
                        inp.get('type', ''),
                        page_url
                    ]
                    ws_elements.append(row)
        
        # Auto-adjust column widths
        for col in ws_elements.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_elements.column_dimensions[col_letter].width = adjusted_width
        
        # Save workbook
        wb.save(excel_file)
        return excel_file
    
    def _get_action_description(self, action: Dict[str, Any]) -> str:
        """Get human-readable description for an action"""
        action_type = action.get('action', '')
        
        if action_type == 'navigate':
            return f"Navigate to {action.get('url', '')}"
        elif action_type == 'page_capture':
            buttons_count = len(action.get('buttons', []))
            inputs_count = len(action.get('inputs', []))
            return f"Page capture: {buttons_count} buttons, {inputs_count} inputs found"
        elif action_type == 'navigation':
            return f"Page navigation to {action.get('url', '')}"
        else:
            return "Unknown action"
    
    def cleanup(self):
        """Clean up browser resources"""
        try:
            if self.page:
                self.page.close()
            if self.context:
                self.context.close()
            if self.browser:
                self.browser.close()
            if self.playwright:
                self.playwright.stop()
        except Exception as e:
            print(f"Error during cleanup: {str(e)}")
    
    def run(self):
        """Main execution method"""
        try:
            # Step 1: Get browser choice
            browser_type, mode = self.get_browser_choice()
            
            # Step 2: Open browser
            self.open_browser(browser_type, mode)
            
            # Step 3: Get URL
            url = self.get_url()
            
            # Step 4: Navigate to URL
            self.navigate_to_url(url)
            
            # Step 5: Start recording
            self.start_recording()
            
            # Step 6: Generate test cases
            self.generate_test_cases()
            
        except Exception as e:
            print(f"\nError: {str(e)}")
            import traceback
            traceback.print_exc()
        finally:
            self.cleanup()


def main():
    """Entry point"""
    print("\n" + "="*60)
    print("Welcome to Test Capture Tool")
    print("="*60)
    print("\nThis tool will help you:")
    print("  1. Open a browser of your choice")
    print("  2. Navigate to a URL")
    print("  3. Record your interactions")
    print("  4. Generate test cases automatically")
    print("="*60)
    
    tool = TestCaptureTool()
    tool.run()


if __name__ == "__main__":
    main()
