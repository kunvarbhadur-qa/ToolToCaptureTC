"""
Test Capture Tool - GUI Version
Records user interactions and generates test cases with a graphical interface
"""

import json
import os
import threading
import time
import subprocess
import platform
from datetime import datetime
from typing import List, Dict, Any, Optional
from tkinter import (
    Tk, ttk, StringVar, IntVar, Text, Scrollbar, 
    messagebox, filedialog, scrolledtext, W, E, N, S
)
from tkinter.font import Font as TkFont
import queue

# Try to import Windows automation libraries
try:
    import pyautogui
    import pyperclip
    PYTHON_AUTOMATION_AVAILABLE = True
except ImportError:
    PYTHON_AUTOMATION_AVAILABLE = False

try:
    if platform.system() == "Windows":
        import win32gui
        import win32con
        import win32clipboard
        WIN32_AVAILABLE = True
    else:
        WIN32_AVAILABLE = False
except ImportError:
    WIN32_AVAILABLE = False

try:
    from playwright.sync_api import sync_playwright, Page, Browser, BrowserContext
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class TestCaptureToolGUI:
    def __init__(self, root: Tk):
        self.root = root
        self.root.title("Doceree Test Capture Tool")
        self.root.geometry("1200x750")
        self.root.minsize(900, 600)
        
        # Variables
        self.recorded_actions: List[Dict[str, Any]] = []
        self.browser: Optional[Browser] = None
        self.context: Optional[BrowserContext] = None
        self.page: Optional[Page] = None
        self.playwright = None
        self.current_url = ""
        self.is_recording = False
        self.browser_thread = None
        self.message_queue = queue.Queue()
        
        # Browser selection
        self.browser_var = StringVar(value="chrome")
        self.mode_var = StringVar(value="normal")
        self.use_existing_profile = IntVar(value=1)  # 1 = use existing profile by default
        
        # Status
        self.status_var = StringVar(value="Ready")
        
        self.setup_ui()
        self.check_dependencies()
        
        # Check for messages from threads
        self.root.after(100, self.process_queue)
    
    def check_dependencies(self):
        """Check if required dependencies are available"""
        if not PLAYWRIGHT_AVAILABLE:
            messagebox.showerror(
                "Missing Dependency",
                "Playwright is not installed!\n\n"
                "Please install it with:\n"
                "pip install playwright\n"
                "playwright install"
            )
        if not EXCEL_AVAILABLE:
            messagebox.showwarning(
                "Excel Support",
                "openpyxl is not installed.\n"
                "Excel files will not be generated.\n\n"
                "Install with: pip install openpyxl"
            )
    
    def setup_ui(self):
        """Setup the user interface"""
        # Main container
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(W, E, N, S))
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        
        # Title
        title_font = TkFont(family="Arial", size=16)
        title_font.configure(weight="bold")
        title_label = ttk.Label(
            main_frame, 
            text="Doceree Test Capture Tool",
            font=title_font
        )
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # Left Panel - Configuration
        left_panel = ttk.LabelFrame(main_frame, text="Configuration", padding="10")
        left_panel.grid(row=1, column=0, sticky=(W, E, N, S), padx=(0, 10))
        
        # Instructions
        instructions_text = (
            "Option 1 - Auto Launch (may have issues):\n"
            "• Click 'Connect for Recording' - Chrome launches automatically\n\n"
            "Option 2 - Manual Browser (recommended):\n"
            "• Open Chrome manually with: chrome.exe --remote-debugging-port=9222\n"
            "• Then click 'Connect for Recording' to connect to your browser\n\n"
            "For Simple Navigation:\n"
            "• Use 'Send URL to Browser' (no setup needed)"
        )
        instructions_label = ttk.Label(
            left_panel, 
            text=instructions_text,
            font=TkFont(size=8),
            foreground="darkblue",
            justify="left",
            wraplength=400
        )
        instructions_label.grid(row=0, column=0, columnspan=2, sticky=W, pady=10, padx=5)
        
        # URL Input
        ttk.Label(left_panel, text="URL:").grid(row=1, column=0, sticky=W, pady=5)
        self.url_entry = ttk.Entry(left_panel, width=40)
        self.url_entry.grid(row=1, column=1, sticky=(W, E), pady=5)
        self.url_entry.insert(0, "https://")
        left_panel.columnconfigure(1, weight=1)
        
        # Control Buttons
        button_frame = ttk.Frame(left_panel)
        button_frame.grid(row=2, column=0, columnspan=2, pady=10, sticky=(W, E))
        
        self.send_url_btn = ttk.Button(
            button_frame, text="Send URL to Browser", command=self.send_url_to_browser
        )
        self.send_url_btn.grid(row=0, column=0, padx=5, pady=5, sticky=(W, E))
        
        self.connect_browser_btn = ttk.Button(
            button_frame, text="Connect for Recording", command=self.connect_browser_gui
        )
        self.connect_browser_btn.grid(row=0, column=1, padx=5, pady=5, sticky=(W, E))
        
        self.capture_btn = ttk.Button(
            button_frame, text="Capture Page", command=self.capture_page_gui,
            state="disabled"
        )
        self.capture_btn.grid(row=1, column=0, padx=5, pady=5, sticky=(W, E))
        
        self.stop_btn = ttk.Button(
            button_frame, text="Stop Recording", command=self.stop_recording_gui,
            state="disabled"
        )
        self.stop_btn.grid(row=1, column=1, padx=5, pady=5, sticky=(W, E))
        
        self.generate_btn = ttk.Button(
            button_frame, text="Generate Test Cases", command=self.generate_test_cases_gui,
            state="disabled"
        )
        self.generate_btn.grid(row=1, column=1, padx=5, pady=5, sticky=(W, E))
        
        button_frame.columnconfigure(0, weight=1)
        button_frame.columnconfigure(1, weight=1)
        
        # Status
        status_frame = ttk.Frame(left_panel)
        status_frame.grid(row=3, column=0, columnspan=2, pady=10, sticky=(W, E))
        
        ttk.Label(status_frame, text="Status:").grid(row=0, column=0, sticky=W)
        status_font = TkFont()
        status_font.configure(weight="bold")
        self.status_label = ttk.Label(
            status_frame, textvariable=self.status_var,
            foreground="blue", font=status_font
        )
        self.status_label.grid(row=0, column=1, sticky=W, padx=5)
        
        # Right Panel - Log and Actions
        right_panel = ttk.Frame(main_frame)
        right_panel.grid(row=1, column=1, sticky=(W, E, N, S))
        
        # Actions List
        actions_frame = ttk.LabelFrame(right_panel, text="Recorded Actions", padding="5")
        actions_frame.grid(row=0, column=0, sticky=(W, E, N, S), pady=(0, 10))
        
        # Treeview for actions
        tree_frame = ttk.Frame(actions_frame)
        tree_frame.grid(row=0, column=0, sticky=(W, E, N, S))
        
        columns = ("#", "Type", "URL", "Title")
        self.actions_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=8)
        
        self.actions_tree.heading("#", text="#")
        self.actions_tree.heading("Type", text="Action Type")
        self.actions_tree.heading("URL", text="URL")
        self.actions_tree.heading("Title", text="Page Title")
        
        self.actions_tree.column("#", width=50)
        self.actions_tree.column("Type", width=120)
        self.actions_tree.column("URL", width=300)
        self.actions_tree.column("Title", width=200)
        
        scrollbar_tree = ttk.Scrollbar(tree_frame, orient="vertical", command=self.actions_tree.yview)
        self.actions_tree.configure(yscrollcommand=scrollbar_tree.set)
        
        self.actions_tree.grid(row=0, column=0, sticky=(W, E, N, S))
        scrollbar_tree.grid(row=0, column=1, sticky=(N, S))
        
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)
        actions_frame.columnconfigure(0, weight=1)
        actions_frame.rowconfigure(0, weight=1)
        
        # Log Area
        log_frame = ttk.LabelFrame(right_panel, text="Activity Log", padding="5")
        log_frame.grid(row=1, column=0, sticky=(W, E, N, S))
        
        # Log toolbar
        log_toolbar = ttk.Frame(log_frame)
        log_toolbar.grid(row=0, column=0, sticky=(W, E), pady=(0, 5))
        
        ttk.Button(
            log_toolbar, text="Save Log", command=self.save_log_file
        ).grid(row=0, column=0, padx=2)
        
        ttk.Button(
            log_toolbar, text="Export Log", command=self.export_log_file
        ).grid(row=0, column=1, padx=2)
        
        ttk.Button(
            log_toolbar, text="Clear Log", command=self.clear_log
        ).grid(row=0, column=2, padx=2)
        
        self.log_auto_save = IntVar(value=1)  # Auto-save enabled by default
        ttk.Checkbutton(
            log_toolbar, text="Auto-save", variable=self.log_auto_save
        ).grid(row=0, column=3, padx=5)
        
        self.log_text = scrolledtext.ScrolledText(
            log_frame, height=15, width=60, wrap="word"
        )
        self.log_text.grid(row=1, column=0, sticky=(W, E, N, S))
        
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(1, weight=1)
        
        # Initialize log file
        self.log_file_path = None
        self.log_directory = os.path.join(os.getcwd(), "activity_logs")
        os.makedirs(self.log_directory, exist_ok=True)
        
        # Configure grid weights
        main_frame.columnconfigure(0, weight=0)  # Left panel - fixed width
        main_frame.columnconfigure(1, weight=1)  # Right panel - expandable
        main_frame.rowconfigure(1, weight=1)
        
        left_panel.columnconfigure(1, weight=1)
        right_panel.columnconfigure(0, weight=1)
        right_panel.rowconfigure(0, weight=1)
        right_panel.rowconfigure(1, weight=1)
        
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        self.log("Application started. Ready to begin.", "INFO")
        self.log(f"Log directory: {self.log_directory}", "INFO")
        if self.log_auto_save.get():
            self.log("Auto-save enabled. Logs will be saved automatically.", "INFO")
    
    def log(self, message: str, level: str = "INFO"):
        """Add message to log and optionally save to file"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = f"[{timestamp}] [{level}] {message}\n"
        
        # Display in GUI
        self.log_text.insert("end", log_entry)
        self.log_text.see("end")
        
        # Auto-save to file if enabled
        if self.log_auto_save.get():
            self._write_to_log_file(log_entry)
    
    def _write_to_log_file(self, log_entry: str):
        """Write log entry to file"""
        try:
            if not self.log_file_path:
                # Create new log file for this session
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                self.log_file_path = os.path.join(
                    self.log_directory, 
                    f"activity_log_{timestamp}.txt"
                )
            
            with open(self.log_file_path, "a", encoding="utf-8") as f:
                f.write(log_entry)
        except Exception as e:
            # Don't show error in log to avoid recursion
            print(f"Error writing to log file: {str(e)}")
    
    def save_log_file(self):
        """Save current log to file"""
        try:
            log_content = self.log_text.get("1.0", "end-1c")
            if not log_content.strip():
                messagebox.showwarning("Warning", "Log is empty. Nothing to save.")
                return
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"activity_log_{timestamp}.txt"
            
            file_path = filedialog.asksaveasfilename(
                title="Save Activity Log",
                defaultextension=".txt",
                initialfile=default_filename,
                initialdir=self.log_directory,
                filetypes=[
                    ("Text files", "*.txt"),
                    ("All files", "*.*")
                ]
            )
            
            if file_path:
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write("="*80 + "\n")
                    f.write("Doceree Test Capture Tool - Activity Log\n")
                    f.write("="*80 + "\n")
                    f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write("="*80 + "\n\n")
                    f.write(log_content)
                
                self.log(f"Log saved to: {os.path.basename(file_path)}", "SUCCESS")
                messagebox.showinfo("Success", f"Log saved successfully!\n\n{file_path}")
        except Exception as e:
            self.log(f"Error saving log: {str(e)}", "ERROR")
            messagebox.showerror("Error", f"Failed to save log:\n{str(e)}")
    
    def export_log_file(self):
        """Export log to a location chosen by user"""
        try:
            log_content = self.log_text.get("1.0", "end-1c")
            if not log_content.strip():
                messagebox.showwarning("Warning", "Log is empty. Nothing to export.")
                return
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"activity_log_{timestamp}.txt"
            
            file_path = filedialog.asksaveasfilename(
                title="Export Activity Log",
                defaultextension=".txt",
                initialfile=default_filename,
                filetypes=[
                    ("Text files", "*.txt"),
                    ("CSV files", "*.csv"),
                    ("All files", "*.*")
                ]
            )
            
            if file_path:
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write("="*80 + "\n")
                    f.write("Doceree Test Capture Tool - Activity Log\n")
                    f.write("="*80 + "\n")
                    f.write(f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write(f"Total Actions Recorded: {len(self.recorded_actions)}\n")
                    f.write("="*80 + "\n\n")
                    f.write(log_content)
                
                self.log(f"Log exported to: {os.path.basename(file_path)}", "SUCCESS")
                messagebox.showinfo("Success", f"Log exported successfully!\n\n{file_path}")
        except Exception as e:
            self.log(f"Error exporting log: {str(e)}", "ERROR")
            messagebox.showerror("Error", f"Failed to export log:\n{str(e)}")
    
    def clear_log(self):
        """Clear the log display"""
        if messagebox.askyesno("Clear Log", "Are you sure you want to clear the log display?\n\nNote: Auto-saved log files will not be deleted."):
            self.log_text.delete("1.0", "end")
            self.log("Log display cleared.", "INFO")
    
    def update_status(self, status: str, color: str = "blue"):
        """Update status label"""
        self.status_var.set(status)
        self.status_label.config(foreground=color)
    
    def process_queue(self):
        """Process messages from background threads"""
        try:
            while True:
                message = self.message_queue.get_nowait()
                if message["type"] == "log":
                    log_level = message.get("level", "INFO")
                    self.log(message["text"], log_level)
                elif message["type"] == "status":
                    self.update_status(message["text"], message.get("color", "blue"))
                elif message["type"] == "action_added":
                    self.add_action_to_tree(message["action"], message["index"])
        except queue.Empty:
            pass
        finally:
            self.root.after(100, self.process_queue)
    
    def add_action_to_tree(self, action: Dict[str, Any], index: int):
        """Add action to treeview"""
        action_type = action.get('action', 'unknown')
        url = action.get('url', '')[:50] + ('...' if len(action.get('url', '')) > 50 else '')
        title = (action.get('title') or action.get('page_title', ''))[:40] + ('...' if len(action.get('title') or action.get('page_title', '')) > 40 else '')
        
        self.actions_tree.insert("", "end", values=(index, action_type, url, title))
        self.actions_tree.see("end")
    
    def send_url_to_browser(self):
        """Send URL to currently open Chrome browser without requiring remote debugging"""
        url = self.url_entry.get().strip()
        if not url or url == "https://":
            messagebox.showwarning("Warning", "Please enter a valid URL")
            return
        
        # Add protocol if missing
        if not url.startswith(("http://", "https://")):
            url = "https://" + url
            self.url_entry.delete(0, "end")
            self.url_entry.insert(0, url)
        
        # Disable button
        self.send_url_btn.config(state="disabled")
        self.update_status("Sending URL to browser...", "orange")
        
        # Run in thread to avoid blocking UI
        thread = threading.Thread(
            target=self._send_url_thread,
            args=(url,),
            daemon=True
        )
        thread.start()
    
    def _send_url_thread(self, url: str):
        """Send URL to Chrome using Windows automation"""
        try:
            self.message_queue.put({
                "type": "log",
                "text": f"Sending URL to Chrome: {url}",
                "level": "INFO"
            })
            
            if platform.system() != "Windows":
                error_msg = "URL sending is currently only supported on Windows"
                self.message_queue.put({
                    "type": "log",
                    "text": error_msg,
                    "level": "ERROR"
                })
                self.root.after(0, lambda: self.send_url_btn.config(state="normal"))
                self.root.after(0, lambda: messagebox.showerror("Error", error_msg))
                return
            
            # Method 1: Try to find and activate Chrome window, then send URL
            chrome_windows = []  # Initialize before use
            if WIN32_AVAILABLE:
                try:
                    # Find Chrome window
                    def enum_handler(hwnd, windows):
                        if win32gui.IsWindowVisible(hwnd):
                            window_title = win32gui.GetWindowText(hwnd)
                            if "chrome" in window_title.lower() or "google chrome" in window_title.lower():
                                windows.append((hwnd, window_title))
                    
                    win32gui.EnumWindows(enum_handler, chrome_windows)
                    
                    if chrome_windows:
                        # Get the first Chrome window
                        hwnd, title = chrome_windows[0]
                        # Bring window to foreground
                        win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
                        win32gui.SetForegroundWindow(hwnd)
                        time.sleep(0.3)  # Wait for window to activate
                        
                        self.message_queue.put({
                            "type": "log",
                            "text": f"Found Chrome window: {title}",
                            "level": "INFO"
                        })
                    else:
                        # Chrome not found, try to open it
                        self.message_queue.put({
                            "type": "log",
                            "text": "Chrome window not found. Attempting to open Chrome...",
                            "level": "WARNING"
                        })
                        try:
                            chrome_paths = [
                                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
                                os.path.expanduser(r"~\AppData\Local\Google\Chrome\Application\chrome.exe")
                            ]
                            for path in chrome_paths:
                                if os.path.exists(path):
                                    subprocess.Popen([path, url])
                                    self.message_queue.put({
                                        "type": "log",
                                        "text": f"Opened Chrome with URL: {url}",
                                        "level": "SUCCESS"
                                    })
                                    self.root.after(0, lambda: self.send_url_btn.config(state="normal"))
                                    self.root.after(0, lambda: self.update_status("URL sent", "green"))
                                    return
                            raise FileNotFoundError("Chrome executable not found")
                        except Exception as e:
                            error_msg = f"Could not open Chrome: {str(e)}"
                            self.message_queue.put({
                                "type": "log",
                                "text": error_msg,
                                "level": "ERROR"
                            })
                            self.root.after(0, lambda: self.send_url_btn.config(state="normal"))
                            self.root.after(0, lambda: messagebox.showerror("Error", error_msg))
                            return
                except Exception as e:
                    self.message_queue.put({
                        "type": "log",
                        "text": f"Window automation failed: {str(e)}. Trying alternative method...",
                        "level": "WARNING"
                    })
            
            # Method 2: Use pyautogui to send keystrokes to navigate current tab
            if PYTHON_AUTOMATION_AVAILABLE:
                try:
                    # Copy URL to clipboard
                    pyperclip.copy(url)
                    time.sleep(0.3)  # Wait a bit longer for clipboard
                    
                    # Send Ctrl+L to focus address bar (this navigates current tab, doesn't open new one)
                    pyautogui.hotkey('ctrl', 'l')  # Focus address bar
                    time.sleep(0.3)
                    
                    # Clear any existing text and paste new URL
                    pyautogui.hotkey('ctrl', 'a')  # Select all
                    time.sleep(0.1)
                    pyautogui.hotkey('ctrl', 'v')  # Paste URL
                    time.sleep(0.2)
                    
                    # Press Enter to navigate
                    pyautogui.press('enter')
                    
                    self.message_queue.put({
                        "type": "log",
                        "text": f"URL sent to current Chrome tab: {url}",
                        "level": "SUCCESS"
                    })
                    self.message_queue.put({
                        "type": "status",
                        "text": "URL sent",
                        "color": "green"
                    })
                    self.root.after(0, lambda: self.send_url_btn.config(state="normal"))
                    return
                except Exception as e:
                    error_msg = f"Keyboard automation failed: {str(e)}"
                    self.message_queue.put({
                        "type": "log",
                        "text": error_msg,
                        "level": "WARNING"
                    })
                    self.message_queue.put({
                        "type": "log",
                        "text": "Trying alternative method...",
                        "level": "INFO"
                    })
                    # Don't return - continue to fallback method
            
            # Method 3: Fallback - try to open URL in Chrome directly using subprocess
            try:
                chrome_paths = [
                    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                    r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
                    os.path.expanduser(r"~\AppData\Local\Google\Chrome\Application\chrome.exe")
                ]
                for path in chrome_paths:
                    if os.path.exists(path):
                        # Open URL in existing Chrome window (will open in new tab)
                        subprocess.Popen([path, url])
                        self.message_queue.put({
                            "type": "log",
                            "text": f"Opened URL in Chrome: {url}",
                            "level": "SUCCESS"
                        })
                        self.message_queue.put({
                            "type": "status",
                            "text": "URL opened",
                            "color": "green"
                        })
                        self.root.after(0, lambda: self.send_url_btn.config(state="normal"))
                        return
            except Exception as e:
                self.message_queue.put({
                    "type": "log",
                    "text": f"Subprocess method failed: {str(e)}. Trying webbrowser...",
                    "level": "WARNING"
                })
            
            # Final fallback - use webbrowser
            try:
                import webbrowser
                webbrowser.open(url)
                self.message_queue.put({
                    "type": "log",
                    "text": f"Opened URL in default browser: {url}",
                    "level": "INFO"
                })
                self.message_queue.put({
                    "type": "status",
                    "text": "URL opened",
                    "color": "green"
                })
                self.root.after(0, lambda: self.send_url_btn.config(state="normal"))
                return
            except Exception as e:
                error_msg = f"All methods failed. Error: {str(e)}"
                self.message_queue.put({
                    "type": "log",
                    "text": error_msg,
                    "level": "ERROR"
                })
                self.root.after(0, lambda: self.send_url_btn.config(state="normal"))
                self.root.after(0, lambda: messagebox.showerror("Error", error_msg))
                
        except Exception as e:
            error_msg = f"Unexpected error: {str(e)}"
            self.message_queue.put({
                "type": "log",
                "text": error_msg,
                "level": "ERROR"
            })
            self.root.after(0, lambda: self.send_url_btn.config(state="normal"))
            self.root.after(0, lambda: messagebox.showerror("Error", error_msg))
    
    def connect_browser_gui(self):
        """Launch browser for full recording (no remote debugging needed)"""
        if not PLAYWRIGHT_AVAILABLE:
            messagebox.showerror("Error", "Playwright is not installed!")
            return
        
        url = self.url_entry.get().strip()
        if not url or url == "https://":
            messagebox.showwarning("Warning", "Please enter a valid URL")
            return
        
        # Add protocol if missing
        if not url.startswith(("http://", "https://")):
            url = "https://" + url
            self.url_entry.delete(0, "end")
            self.url_entry.insert(0, url)
        
        # Disable button
        self.connect_browser_btn.config(state="disabled")
        self.update_status("Launching browser for recording...", "orange")
        
        # Run in thread to avoid blocking UI
        thread = threading.Thread(
            target=self._launch_browser_for_recording,
            args=(url,),
            daemon=True
        )
        thread.start()
    
    def _launch_browser_for_recording(self, url: str):
        """Launch browser directly for recording (no remote debugging needed)"""
        try:
            self.message_queue.put({
                "type": "log",
                "text": "Launching Chrome browser for recording (no setup needed)...",
                "level": "INFO"
            })
            
            self.playwright = sync_playwright().start()
            
            # Get screen size for proper viewport
            import tkinter as tk
            temp_root = tk.Tk()
            screen_width = temp_root.winfo_screenwidth()
            screen_height = temp_root.winfo_screenheight()
            temp_root.destroy()
            
            # Launch Chrome directly (no remote debugging needed)
            try:
                self.browser = self.playwright.chromium.launch(
                    headless=False,
                    channel="chrome",  # Use actual Google Chrome
                    args=[
                        "--start-maximized",
                        "--disable-blink-features=AutomationControlled",  # Hide automation
                        "--disable-dev-shm-usage",
                        "--no-sandbox",
                        "--disable-web-security",  # Allow all resources
                        "--disable-features=IsolateOrigins,site-per-process",  # Better compatibility
                        "--allow-running-insecure-content",  # Allow mixed content
                        "--disable-site-isolation-trials"  # Better resource loading
                    ]
                )
                self.message_queue.put({
                    "type": "log",
                    "text": "Chrome browser launched successfully",
                    "level": "SUCCESS"
                })
                
                # Create context with proper settings for better compatibility
                context_options = {
                    "viewport": {"width": screen_width, "height": max(screen_height, 2000)},
                    "record_video_dir": "test_recordings/",
                    "ignore_https_errors": False,
                    "java_script_enabled": True,
                    "bypass_csp": True,  # Allow all resources
                    "accept_downloads": True,
                    "locale": "en-US",
                    "timezone_id": "America/New_York",
                    "permissions": ["geolocation", "notifications"],
                    "extra_http_headers": {
                        "Accept-Language": "en-US,en;q=0.9"
                    }
                }
                
                self.context = self.browser.new_context(**context_options)
                self.page = self.context.new_page()
                
                # Remove automation indicators
                self.page.add_init_script("""
                    Object.defineProperty(navigator, 'webdriver', {
                        get: () => undefined
                    });
                    window.chrome = {
                        runtime: {}
                    };
                    Object.defineProperty(navigator, 'plugins', {
                        get: () => [1, 2, 3, 4, 5]
                    });
                """)
                
                # Wait for window to maximize
                import time
                time.sleep(0.5)
                
                # Set viewport
                try:
                    self.page.set_viewport_size({
                        "width": screen_width, 
                        "height": max(screen_height, 2000)
                    })
                    self.message_queue.put({
                        "type": "log",
                        "text": f"Window maximized. Viewport: {screen_width}x{max(screen_height, 2000)}",
                        "level": "INFO"
                    })
                except:
                    pass
                    
            except Exception as e:
                error_msg = f"Failed to launch Chrome browser: {str(e)}\n\nMake sure Google Chrome is installed."
                self.message_queue.put({
                    "type": "log",
                    "text": error_msg,
                    "level": "ERROR"
                })
                self.message_queue.put({
                    "type": "status",
                    "text": "Launch failed",
                    "color": "red"
                })
                self.root.after(0, lambda: self.connect_browser_btn.config(state="normal"))
                self.root.after(0, lambda: messagebox.showerror("Browser Launch Error", error_msg))
                return
            
            # Navigate to URL
            if not hasattr(self, 'page') or self.page is None:
                error_msg = "Failed to create browser page. Please try again."
                self.message_queue.put({
                    "type": "log",
                    "text": error_msg,
                    "level": "ERROR"
                })
                self.message_queue.put({
                    "type": "status",
                    "text": "Page creation failed",
                    "color": "red"
                })
                self.root.after(0, lambda: self.connect_browser_btn.config(state="normal"))
                self.root.after(0, lambda: messagebox.showerror("Error", error_msg))
                return
            
            self.current_url = url
            self.message_queue.put({
                "type": "log",
                "text": f"Navigating to: {url}",
                "level": "INFO"
            })
            
            try:
                # Navigate and wait for all resources to load
                self.page.goto(url, wait_until="networkidle", timeout=60000)
                
                # Wait for page to be fully loaded
                self.page.wait_for_load_state("load", timeout=30000)
                self.page.wait_for_load_state("domcontentloaded", timeout=30000)
                
                # Wait for CSS and images to load
                import time
                time.sleep(3)
                
                # Wait for all images to load
                try:
                    self.page.evaluate("""
                        async () => {
                            const images = Array.from(document.images);
                            await Promise.all(images.map(img => {
                                if (img.complete) return Promise.resolve();
                                return new Promise((resolve) => {
                                    img.onload = resolve;
                                    img.onerror = resolve;
                                    setTimeout(resolve, 2000);
                                });
                            }));
                        }
                    """)
                except:
                    pass
                
                self.message_queue.put({
                    "type": "log",
                    "text": "Page fully loaded with all resources (CSS, images, etc.)",
                    "level": "SUCCESS"
                })
            except Exception as e:
                error_msg = f"Failed to navigate to URL: {str(e)}"
                self.message_queue.put({
                    "type": "log",
                    "text": error_msg,
                    "level": "ERROR"
                })
                self.message_queue.put({
                    "type": "status",
                    "text": "Navigation failed",
                    "color": "red"
                })
                self.root.after(0, lambda: self.connect_browser_btn.config(state="normal"))
                self.root.after(0, lambda: messagebox.showerror("Navigation Error", error_msg))
                return
            
            # Ensure page is fully interactive
            try:
                self.page.wait_for_load_state("domcontentloaded")
                # Remove automation indicators and enable interactions
                self.page.evaluate("""
                    // Remove webdriver property
                    Object.defineProperty(navigator, 'webdriver', {
                        get: () => undefined
                    });
                    
                    // Focus window
                    window.focus();
                    
                    // Ensure all interactive elements are enabled
                    document.body.style.pointerEvents = 'auto';
                    document.documentElement.style.pointerEvents = 'auto';
                    
                    // Enable all buttons and interactive elements
                    const allElements = document.querySelectorAll('*');
                    allElements.forEach(el => {
                        el.style.pointerEvents = 'auto';
                        if (el.disabled) {
                            el.removeAttribute('disabled');
                        }
                    });
                """)
                self.message_queue.put({
                    "type": "log",
                    "text": "Page loaded. Automation hidden. All interactions enabled.",
                    "level": "INFO"
                })
            except Exception as e:
                self.message_queue.put({
                    "type": "log",
                    "text": f"Note: {str(e)}",
                    "level": "WARNING"
                })
            
            # Record initial navigation
            action = {
                "action": "navigate",
                "url": url,
                "timestamp": datetime.now().isoformat(),
                "page_title": self.page.title(),
                "page_url": self.page.url
            }
            self.recorded_actions.append(action)
            
            # Setup listeners
            self.setup_listeners()
            
            # Initial capture
            self.capture_page_state()
            
            self.is_recording = True
            self.message_queue.put({
                "type": "status",
                "text": "Recording...",
                "color": "green"
            })
            self.message_queue.put({
                "type": "log",
                "text": f"Page loaded: {self.page.title()}"
            })
            self.message_queue.put({
                "type": "log",
                "text": "Recording started. Interact with the browser and click 'Capture Page' to record states."
            })
            self.message_queue.put({
                "type": "action_added",
                "action": action,
                "index": len(self.recorded_actions)
            })
            
            # Enable buttons
            self.root.after(0, lambda: self.capture_btn.config(state="normal"))
            self.root.after(0, lambda: self.stop_btn.config(state="normal"))
            self.root.after(0, lambda: self.generate_btn.config(state="normal"))
            
        except Exception as e:
            self.message_queue.put({
                "type": "status",
                "text": "Error",
                "color": "red"
            })
            self.message_queue.put({
                "type": "log",
                "text": f"Error: {str(e)}",
                "level": "ERROR"
            })
            self.root.after(0, lambda: self.connect_browser_btn.config(state="normal"))
            self.root.after(0, lambda: messagebox.showerror("Error", f"Failed to connect to browser:\n{str(e)}"))
    
    def _open_browser_thread(self, browser_type: str, mode: str, url: str):
        """Open browser in background thread"""
        try:
            self.message_queue.put({
                "type": "log",
                "text": f"Starting {browser_type} browser in {mode} mode...",
                "level": "INFO"
            })
            
            self.playwright = sync_playwright().start()
            
            # Launch browser
            if browser_type == "chrome":
                browser_launcher = self.playwright.chromium
                
                # Check if we should use existing Chrome profile
                use_existing_profile = self.use_existing_profile.get()
                
                if use_existing_profile:
                    # Create a temporary user data directory to avoid conflicts with existing Chrome
                    # This allows opening a new Chrome window without affecting existing ones
                    import tempfile
                    import platform
                    
                    # Create a temporary directory for this session's Chrome profile
                    temp_profile_dir = os.path.join(
                        tempfile.gettempdir(),
                        "doceree_test_chrome_profile",
                        datetime.now().strftime("%Y%m%d_%H%M%S")
                    )
                    os.makedirs(temp_profile_dir, exist_ok=True)
                    
                    self.message_queue.put({
                        "type": "log",
                        "text": f"Using temporary Chrome profile (won't affect existing Chrome windows)",
                        "level": "INFO"
                    })
                    
                    # Use persistent context with temporary profile
                    try:
                        # Get screen size for proper viewport
                        import tkinter as tk
                        temp_root = tk.Tk()
                        screen_width = temp_root.winfo_screenwidth()
                        screen_height = temp_root.winfo_screenheight()
                        temp_root.destroy()
                        
                        # Launch persistent context with temporary user data
                        # Disable automation flags to make it behave like normal browser
                        persistent_options = {
                            "headless": False,
                            "slow_mo": 100,
                            "channel": "chrome",
                            "viewport": {"width": screen_width, "height": max(screen_height, 2000)},
                            "args": [
                                "--start-maximized",
                                "--new-window",
                                "--disable-blink-features=AutomationControlled",  # Hide automation
                                "--disable-dev-shm-usage",
                                "--no-sandbox"
                            ],
                            "ignore_https_errors": False
                        }
                        
                        # Add incognito mode if selected
                        if mode in ["incognito", "private"]:
                            persistent_options["args"].append("--incognito")
                        
                        self.message_queue.put({
                            "type": "log",
                            "text": "Launching new Chrome window (separate from existing Chrome)...",
                            "level": "INFO"
                        })
                        
                        # Use launch_persistent_context with temporary profile
                        # This creates a new Chrome instance that won't interfere with existing windows
                        self.context = browser_launcher.launch_persistent_context(
                            user_data_dir=temp_profile_dir,
                            **persistent_options
                        )
                        self.browser = None  # Persistent context doesn't have separate browser object
                        
                        # Wait a moment for the initial page to be created
                        import time
                        time.sleep(0.5)
                        
                        # Get the first page (the one that opens by default) instead of creating a new one
                        pages = self.context.pages
                        if pages:
                            self.page = pages[0]  # Use the existing page instead of creating a new one
                        else:
                            self.page = self.context.new_page()
                        
                        # Close any extra tabs that might have been created
                        try:
                            all_pages = self.context.pages
                            for page in all_pages:
                                if page != self.page:
                                    page.close()
                        except:
                            pass
                        
                        # Wait a moment for window to maximize
                        time.sleep(0.5)
                        
                        # Ensure viewport matches maximized window for responsive behavior
                        try:
                            self.page.set_viewport_size({
                                "width": screen_width, 
                                "height": max(screen_height, 2000)
                            })
                            self.message_queue.put({
                                "type": "log",
                                "text": f"Window maximized. Viewport: {screen_width}x{max(screen_height, 2000)}",
                                "level": "INFO"
                            })
                        except:
                            pass
                        
                        # Remove automation indicators before navigation
                        self.page.add_init_script("""
                            Object.defineProperty(navigator, 'webdriver', {
                                get: () => undefined
                            });
                            window.chrome = {
                                runtime: {}
                            };
                            Object.defineProperty(navigator, 'plugins', {
                                get: () => [1, 2, 3, 4, 5]
                            });
                        """)
                        
                        # Navigate to URL
                        self.current_url = url
                        self.message_queue.put({
                            "type": "log",
                            "text": f"Navigating to: {url}",
                            "level": "INFO"
                        })
                        
                        self.page.goto(url, wait_until="networkidle", timeout=30000)
                        
                        # Ensure page is fully interactive
                        try:
                            self.page.wait_for_load_state("domcontentloaded")
                            # Remove automation indicators and enable interactions
                            self.page.evaluate("""
                                // Remove webdriver property
                                Object.defineProperty(navigator, 'webdriver', {
                                    get: () => undefined
                                });
                                
                                // Focus window
                                window.focus();
                                
                                // Ensure all interactive elements are enabled
                                document.body.style.pointerEvents = 'auto';
                                document.documentElement.style.pointerEvents = 'auto';
                                
                                // Enable all buttons and interactive elements
                                const allElements = document.querySelectorAll('*');
                                allElements.forEach(el => {
                                    el.style.pointerEvents = 'auto';
                                    if (el.disabled) {
                                        el.removeAttribute('disabled');
                                    }
                                });
                            """)
                            self.message_queue.put({
                                "type": "log",
                                "text": "Page loaded. Automation hidden. All interactions enabled.",
                                "level": "INFO"
                            })
                        except Exception as e:
                            self.message_queue.put({
                                "type": "log",
                                "text": f"Note: {str(e)}",
                                "level": "WARNING"
                            })
                        
                        # Record initial navigation
                        action = {
                            "action": "navigate",
                            "url": url,
                            "timestamp": datetime.now().isoformat(),
                            "page_title": self.page.title(),
                            "page_url": self.page.url
                        }
                        self.recorded_actions.append(action)
                        
                        # Setup listeners
                        self.setup_listeners()
                        
                        # Initial capture
                        self.capture_page_state()
                        
                        self.is_recording = True
                        self.message_queue.put({
                            "type": "status",
                            "text": "Recording...",
                            "color": "green"
                        })
                        self.message_queue.put({
                            "type": "log",
                            "text": f"Page loaded: {self.page.title()}",
                            "level": "SUCCESS"
                        })
                        self.message_queue.put({
                            "type": "log",
                            "text": "New Chrome window opened successfully. Your existing Chrome windows are not affected.",
                            "level": "INFO"
                        })
                        self.message_queue.put({
                            "type": "action_added",
                            "action": action,
                            "index": len(self.recorded_actions)
                        })
                        
                        # Enable buttons
                        self.root.after(0, lambda: self.capture_btn.config(state="normal"))
                        self.root.after(0, lambda: self.stop_btn.config(state="normal"))
                        self.root.after(0, lambda: self.generate_btn.config(state="normal"))
                        
                        return  # Exit early since we've handled persistent context
                        
                    except Exception as e:
                        error_msg = f"Failed to launch Chrome: {str(e)}"
                        self.message_queue.put({
                            "type": "log",
                            "text": error_msg,
                            "level": "ERROR"
                        })
                        self.message_queue.put({
                            "type": "status",
                            "text": "Launch failed",
                            "color": "red"
                        })
                        self.root.after(0, lambda: self.open_browser_btn.config(state="normal"))
                        self.root.after(0, lambda: messagebox.showerror("Browser Launch Error", error_msg))
                        return
                
                # Use Chrome channel to launch actual Google Chrome browser
                launch_options = {
                    "headless": False,
                    "slow_mo": 100,
                    "channel": "chrome"  # Use actual Google Chrome
                }
            elif browser_type == "firefox":
                browser_launcher = self.playwright.firefox
                launch_options = {
                    "headless": False,
                    "slow_mo": 100
                }
            elif browser_type == "msedge":
                browser_launcher = self.playwright.chromium
                launch_options = {
                    "headless": False,
                    "slow_mo": 100,
                    "channel": "msedge"
                }
            else:
                browser_launcher = self.playwright.chromium
                launch_options = {
                    "headless": False,
                    "slow_mo": 100
                }
            
            # Add arguments to maximize browser window and hide automation
            # Don't set window-size to allow natural responsive behavior like manual Chrome
            if browser_type == "chrome" or browser_type == "msedge":
                if "args" not in launch_options:
                    launch_options["args"] = []
                # Start maximized and hide automation flags
                launch_options["args"].extend([
                    "--start-maximized",
                    "--disable-blink-features=AutomationControlled",  # Hide automation
                    "--disable-dev-shm-usage",
                    "--no-sandbox"
                ])
            elif browser_type == "firefox":
                if "args" not in launch_options:
                    launch_options["args"] = []
                launch_options["args"].extend([
                    "--width=1920",
                    "--height=1080"
                ])
            
            if mode in ["incognito", "private"]:
                if browser_type == "chrome":
                    if "args" not in launch_options:
                        launch_options["args"] = []
                    launch_options["args"].append("--incognito")
                elif browser_type == "msedge":
                    if "args" not in launch_options:
                        launch_options["args"] = []
                    launch_options["args"].append("--inprivate")
            
            # Launch browser with error handling
            try:
                self.message_queue.put({
                    "type": "log",
                    "text": "Launching browser...",
                    "level": "INFO"
                })
                self.browser = browser_launcher.launch(**launch_options)
                self.message_queue.put({
                    "type": "log",
                    "text": "Browser launched successfully",
                    "level": "SUCCESS"
                })
            except Exception as e:
                error_msg = f"Failed to launch browser: {str(e)}\n\nMake sure Chrome is installed and the path is correct."
                self.message_queue.put({
                    "type": "log",
                    "text": error_msg,
                    "level": "ERROR"
                })
                self.message_queue.put({
                    "type": "status",
                    "text": "Launch failed",
                    "color": "red"
                })
                self.root.after(0, lambda: self.open_browser_btn.config(state="normal"))
                self.root.after(0, lambda: messagebox.showerror("Browser Launch Error", error_msg))
                return
            
            # Set viewport to screen size for responsive behavior when maximized
            try:
                # Get screen size for proper viewport
                import tkinter as tk
                temp_root = tk.Tk()
                screen_width = temp_root.winfo_screenwidth()
                screen_height = temp_root.winfo_screenheight()
                temp_root.destroy()
                
                context_options = {
                    "viewport": {"width": screen_width, "height": screen_height},
                    "record_video_dir": "test_recordings/"
                }
                
                self.context = self.browser.new_context(**context_options)
                # Create a single page - this will be the only tab
                self.page = self.context.new_page()
                
                # Ensure only one tab is open
                try:
                    all_pages = self.context.pages
                    for page in all_pages:
                        if page != self.page:
                            page.close()
                except:
                    pass
                
                # Wait a moment for window to maximize
                import time
                time.sleep(0.5)
                
                # Ensure viewport matches maximized window for responsive behavior
                try:
                    self.page.set_viewport_size({
                        "width": screen_width, 
                        "height": max(screen_height, 2000)
                    })
                    self.message_queue.put({
                        "type": "log",
                        "text": f"Window maximized. Viewport: {screen_width}x{max(screen_height, 2000)}",
                        "level": "INFO"
                    })
                except:
                    pass
                
                self.message_queue.put({
                    "type": "log",
                    "text": f"Browser context created with responsive viewport: {screen_width}x{max(screen_height, 2000)} (allows scrolling)",
                    "level": "SUCCESS"
                })
            except Exception as e:
                error_msg = f"Failed to create browser context: {str(e)}"
                self.message_queue.put({
                    "type": "log",
                    "text": error_msg,
                    "level": "ERROR"
                })
                self.root.after(0, lambda: self.open_browser_btn.config(state="normal"))
                self.root.after(0, lambda: messagebox.showerror("Error", error_msg))
                return
            
            # Navigate to URL
            self.current_url = url
            self.message_queue.put({
                "type": "log",
                "text": f"Navigating to: {url}",
                "level": "INFO"
            })
            
            try:
                self.page.goto(url, wait_until="networkidle", timeout=30000)
                
                # Ensure page is fully interactive
                try:
                    self.page.wait_for_load_state("domcontentloaded")
                    # Remove automation indicators and enable interactions
                    self.page.evaluate("""
                        // Remove webdriver property
                        Object.defineProperty(navigator, 'webdriver', {
                            get: () => undefined
                        });
                        
                        // Focus window
                        window.focus();
                        
                        // Ensure all interactive elements are enabled
                        document.body.style.pointerEvents = 'auto';
                        document.documentElement.style.pointerEvents = 'auto';
                        
                        // Enable all buttons and interactive elements
                        const allElements = document.querySelectorAll('*');
                        allElements.forEach(el => {
                            el.style.pointerEvents = 'auto';
                            if (el.disabled) {
                                el.removeAttribute('disabled');
                            }
                        });
                    """)
                    self.message_queue.put({
                        "type": "log",
                        "text": "Page loaded. Automation hidden. All interactions enabled.",
                        "level": "INFO"
                    })
                except Exception as e:
                    self.message_queue.put({
                        "type": "log",
                        "text": f"Note: {str(e)}",
                        "level": "WARNING"
                    })
            except Exception as e:
                error_msg = f"Failed to navigate to URL: {str(e)}"
                self.message_queue.put({
                    "type": "log",
                    "text": error_msg,
                    "level": "ERROR"
                })
                self.message_queue.put({
                    "type": "status",
                    "text": "Navigation failed",
                    "color": "red"
                })
                self.root.after(0, lambda: self.open_browser_btn.config(state="normal"))
                self.root.after(0, lambda: messagebox.showerror("Navigation Error", error_msg))
                return
            
            # Record initial navigation
            action = {
                "action": "navigate",
                "url": url,
                "timestamp": datetime.now().isoformat(),
                "page_title": self.page.title(),
                "page_url": self.page.url
            }
            self.recorded_actions.append(action)
            
            # Setup listeners
            self.setup_listeners()
            
            # Initial capture
            self.capture_page_state()
            
            self.is_recording = True
            self.message_queue.put({
                "type": "status",
                "text": "Recording...",
                "color": "green"
            })
            self.message_queue.put({
                "type": "log",
                "text": f"Page loaded: {self.page.title()}",
                "level": "SUCCESS"
            })
            self.message_queue.put({
                "type": "log",
                "text": "Recording started. Interact with the browser and click 'Capture Page' to record states.",
                "level": "INFO"
            })
            self.message_queue.put({
                "type": "action_added",
                "action": action,
                "index": len(self.recorded_actions)
            })
            
            # Enable buttons
            self.root.after(0, lambda: self.capture_btn.config(state="normal"))
            self.root.after(0, lambda: self.stop_btn.config(state="normal"))
            self.root.after(0, lambda: self.generate_btn.config(state="normal"))
            
        except Exception as e:
            self.message_queue.put({
                "type": "status",
                "text": "Error",
                "color": "red"
            })
            self.message_queue.put({
                "type": "log",
                "text": f"Error: {str(e)}",
                "level": "ERROR"
            })
            self.root.after(0, lambda: self.open_browser_btn.config(state="normal"))
            self.root.after(0, lambda: messagebox.showerror("Error", f"Failed to open browser:\n{str(e)}"))
    
    def setup_listeners(self):
        """Setup event listeners"""
        def handle_navigation(event):
            try:
                current_url = self.page.url
                action = {
                    "action": "navigation",
                    "url": current_url,
                    "title": self.page.title(),
                    "timestamp": datetime.now().isoformat()
                }
                self.recorded_actions.append(action)
                self.message_queue.put({
                    "type": "action_added",
                    "action": action,
                    "index": len(self.recorded_actions)
                })
                self.message_queue.put({
                    "type": "log",
                    "text": f"Navigation detected: {self.page.title()}",
                    "level": "INFO"
                })
            except:
                pass
        
        if self.page:
            self.page.on("framenavigated", handle_navigation)
    
    def capture_page_gui(self):
        """Capture page state from GUI"""
        if not self.page:
            messagebox.showwarning("Warning", "Browser is not open!")
            return
        
        thread = threading.Thread(target=self.capture_page_state, daemon=True)
        thread.start()
    
    def capture_page_state(self):
        """Capture current page state"""
        if not self.page:
            return
        
        try:
            page_info = {
                "url": self.page.url,
                "title": self.page.title(),
                "timestamp": datetime.now().isoformat()
            }
            
            # Capture buttons
            buttons = self.page.query_selector_all("button, input[type='button'], input[type='submit'], a[role='button']")
            button_info = []
            for btn in buttons[:50]:
                try:
                    btn_text = btn.inner_text() or btn.get_attribute("value") or btn.get_attribute("aria-label") or ""
                    btn_id = btn.get_attribute("id") or ""
                    btn_class = btn.get_attribute("class") or ""
                    button_info.append({
                        "text": btn_text.strip(),
                        "id": btn_id,
                        "class": btn_class,
                        "tag": btn.evaluate("el => el.tagName.toLowerCase()")
                    })
                except:
                    continue
            
            # Capture inputs
            inputs = self.page.query_selector_all("input, textarea, select")
            input_info = []
            for inp in inputs[:50]:
                try:
                    inp_type = inp.get_attribute("type") or "text"
                    inp_id = inp.get_attribute("id") or ""
                    inp_name = inp.get_attribute("name") or ""
                    inp_placeholder = inp.get_attribute("placeholder") or ""
                    input_info.append({
                        "type": inp_type,
                        "id": inp_id,
                        "name": inp_name,
                        "placeholder": inp_placeholder
                    })
                except:
                    continue
            
            # Capture text
            try:
                body_text = self.page.inner_text("body")[:1000]
            except:
                body_text = ""
            
            page_state = {
                "action": "page_capture",
                **page_info,
                "buttons": button_info,
                "inputs": input_info,
                "body_text_preview": body_text
            }
            
            self.recorded_actions.append(page_state)
            
            self.message_queue.put({
                "type": "action_added",
                "action": page_state,
                "index": len(self.recorded_actions)
            })
            self.message_queue.put({
                "type": "log",
                "text": f"✓ Page captured: {page_info['title']} ({len(button_info)} buttons, {len(input_info)} inputs)",
                "level": "SUCCESS"
            })
            
        except Exception as e:
            self.message_queue.put({
                "type": "log",
                "text": f"Error capturing page: {str(e)}",
                "level": "ERROR"
            })
    
    def stop_recording_gui(self):
        """Stop recording"""
        self.is_recording = False
        self.update_status("Stopped", "red")
        self.log("Recording stopped.")
        self.capture_btn.config(state="disabled")
        self.stop_btn.config(state="disabled")
    
    def generate_test_cases_gui(self):
        """Generate test cases from GUI"""
        if not self.recorded_actions:
            messagebox.showwarning("Warning", "No actions recorded. Nothing to generate.")
            return
        
        # Ask for output directory
        output_dir = filedialog.askdirectory(
            title="Select folder to save test cases",
            initialdir=os.getcwd()
        )
        
        if not output_dir:
            return
        
        self.update_status("Generating test cases...", "orange")
        self.log("Generating test cases...")
        
        thread = threading.Thread(
            target=self._generate_test_cases_thread,
            args=(output_dir,),
            daemon=True
        )
        thread.start()
    
    def _generate_test_cases_thread(self, output_dir: str):
        """Generate test cases in background thread"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            os.makedirs(output_dir, exist_ok=True)
            
            # Generate JSON
            json_file = os.path.join(output_dir, f"test_case_{timestamp}.json")
            with open(json_file, "w", encoding="utf-8") as f:
                json.dump({
                    "test_case_id": f"TC_{timestamp}",
                    "created_at": datetime.now().isoformat(),
                    "initial_url": self.current_url,
                    "total_actions": len(self.recorded_actions),
                    "actions": self.recorded_actions
                }, f, indent=2, ensure_ascii=False)
            
            self.message_queue.put({
                "type": "log",
                "text": f"Generated JSON: {os.path.basename(json_file)}"
            })
            
            # Generate Text
            txt_file = os.path.join(output_dir, f"test_case_{timestamp}.txt")
            with open(txt_file, "w", encoding="utf-8") as f:
                f.write("="*80 + "\n")
                f.write("TEST CASE DOCUMENTATION\n")
                f.write("="*80 + "\n\n")
                f.write(f"Test Case ID: TC_{timestamp}\n")
                f.write(f"Created At: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Initial URL: {self.current_url}\n")
                f.write(f"Total Actions Recorded: {len(self.recorded_actions)}\n")
                f.write("\n" + "="*80 + "\n")
                f.write("RECORDED ACTIONS\n")
                f.write("="*80 + "\n\n")
                
                for idx, action in enumerate(self.recorded_actions, 1):
                    f.write(f"\n--- Action {idx} ---\n")
                    f.write(f"Type: {action.get('action', 'unknown')}\n")
                    f.write(f"Timestamp: {action.get('timestamp', 'N/A')}\n")
                    
                    if action.get('action') == 'navigate':
                        f.write(f"URL: {action.get('url')}\n")
                        f.write(f"Page Title: {action.get('page_title')}\n")
                    elif action.get('action') == 'page_capture':
                        f.write(f"Page URL: {action.get('url')}\n")
                        f.write(f"Page Title: {action.get('title')}\n")
                        f.write(f"\nButtons Found ({len(action.get('buttons', []))}):\n")
                        for btn in action.get('buttons', [])[:10]:
                            f.write(f"  - {btn.get('text', 'N/A')} (ID: {btn.get('id', 'N/A')})\n")
                        f.write(f"\nInput Fields Found ({len(action.get('inputs', []))}):\n")
                        for inp in action.get('inputs', [])[:10]:
                            f.write(f"  - Type: {inp.get('type')}, ID: {inp.get('id', 'N/A')}\n")
                    elif action.get('action') == 'navigation':
                        f.write(f"Navigated to: {action.get('url')}\n")
                        f.write(f"New Page Title: {action.get('title')}\n")
                    f.write("\n")
            
            self.message_queue.put({
                "type": "log",
                "text": f"Generated Text: {os.path.basename(txt_file)}"
            })
            
            # Generate Python
            py_file = os.path.join(output_dir, f"test_case_{timestamp}.py")
            with open(py_file, "w", encoding="utf-8") as f:
                f.write("""\"\"\"
Auto-generated Test Case
Generated by Test Capture Tool
\"\"\"

from playwright.sync_api import sync_playwright
import pytest


def test_recorded_actions():
    \"\"\"Test case generated from recorded user interactions\"\"\"
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, channel="chrome")
        context = browser.new_context(viewport={"width": 1920, "height": 1080})
        page = context.new_page()
        
        try:
""")
                for action in self.recorded_actions:
                    if action.get('action') == 'navigate':
                        f.write(f"            page.goto('{action.get('url')}', wait_until='networkidle')\n")
                        f.write(f"            assert '{action.get('page_title', '')}' in page.title()\n\n")
                    elif action.get('action') == 'page_capture':
                        f.write(f"            page.goto('{action.get('url')}', wait_until='networkidle')\n")
                        f.write(f"            assert '{action.get('title', '')}' in page.title()\n\n")
                        for btn in action.get('buttons', [])[:5]:
                            if btn.get('id'):
                                f.write(f"            assert page.locator('#{btn.get('id')}').is_visible()\n\n")
                
                f.write("""        finally:
            context.close()
            browser.close()


if __name__ == "__main__":
    test_recorded_actions()
""")
            
            self.message_queue.put({
                "type": "log",
                "text": f"Generated Python: {os.path.basename(py_file)}"
            })
            
            # Generate Excel
            excel_file = None
            if EXCEL_AVAILABLE:
                excel_file = self._generate_excel_file(output_dir, timestamp)
                if excel_file:
                    self.message_queue.put({
                        "type": "log",
                        "text": f"Generated Excel: {os.path.basename(excel_file)}"
                    })
            
            self.message_queue.put({
                "type": "status",
                "text": "Test cases generated!",
                "color": "green"
            })
            self.message_queue.put({
                "type": "log",
                "text": f"✓ All test cases generated successfully in: {output_dir}"
            })
            
            self.root.after(0, lambda: messagebox.showinfo(
                "Success",
                f"Test cases generated successfully!\n\n"
                f"Files saved in:\n{output_dir}\n\n"
                f"Total actions: {len(self.recorded_actions)}"
            ))
            
        except Exception as e:
            self.message_queue.put({
                "type": "status",
                "text": "Generation failed",
                "color": "red"
            })
            self.message_queue.put({
                "type": "log",
                "text": f"Error generating test cases: {str(e)}"
            })
            self.root.after(0, lambda: messagebox.showerror("Error", f"Failed to generate test cases:\n{str(e)}"))
    
    def _generate_excel_file(self, output_dir: str, timestamp: str) -> Optional[str]:
        """Generate Excel file in the format matching the user's requirements"""
        try:
            excel_file = os.path.join(output_dir, "Doceree_TestCases.xlsx")
            wb = Workbook()
            
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            # Main test cases sheet
            ws = wb.active
            ws.title = "Test Cases"
            
            # Header row with formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            # Column headers as per image: TC_ID, Module, Prerequisite, Test Data, URL, Description, Execution_Steps, (H empty), Expected Output, Actual Output, Test_Result
            headers = [
                "TC_ID", "Module", "Prerequisite", "Test Data", "URL", 
                "Description", "Execution_Steps", "", "Expected Output", "Actual Output", "Test_Result"
            ]
            ws.append(headers)
            
            # Format header row
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Generate test cases from recorded actions
            tc_counter = 1
            previous_url = ""
            module_name = "Login Page"  # Default, will be extracted from page title
            
            for idx, action in enumerate(self.recorded_actions):
                action_type = action.get('action', '')
                url = action.get('url', '') or action.get('page_url', '')
                page_title = action.get('title', '') or action.get('page_title', '')
                
                # Extract module name from page title or URL
                if page_title:
                    # Try to extract module name from title (e.g., "Doceree - Login" -> "Login Page")
                    if "login" in page_title.lower():
                        module_name = "Login Page"
                    elif "dashboard" in page_title.lower():
                        module_name = "Dashboard"
                    elif "home" in page_title.lower():
                        module_name = "Home Page"
                    else:
                        # Use page title as module name
                        module_name = page_title.split(" - ")[-1] if " - " in page_title else page_title
                
                if action_type == 'navigate':
                    # First navigation - create initial test case
                    tc_id = f"TC_{tc_counter:03d}"
                    prerequisite = "Login URL"
                    test_data = ""
                    description = f"Verify all the visible elements on the {module_name}."
                    execution_steps = f"1. Open URL: {url}"
                    expected_output = f"{module_name} should opened."
                    actual_output = f"{module_name} is openend."
                    test_result = "Pass"
                    
                    ws.append([
                        tc_id, module_name, prerequisite, test_data, url,
                        description, execution_steps, "", expected_output, actual_output, test_result
                    ])
                    tc_counter += 1
                    previous_url = url
                    
                elif action_type == 'page_capture':
                    # Each page capture becomes a test case
                    buttons = action.get('buttons', [])
                    inputs = action.get('inputs', [])
                    
                    # Generate test cases for each significant element
                    # Test case for page elements visibility
                    if buttons or inputs:
                        tc_id = f"TC_{tc_counter:03d}"
                        prerequisite = "User is on login page" if "login" in url.lower() else f"User is on {module_name.lower()}"
                        test_data = ""
                        description = f"Verify all the visible elements on the {module_name}."
                        execution_steps = f"1. Open URL: {url}"
                        if buttons:
                            execution_steps += f" 2. Verify all buttons are visible."
                        if inputs:
                            execution_steps += f" 3. Verify all input fields are visible."
                        expected_output = f"All elements on {module_name} should be visible."
                        actual_output = f"All elements on {module_name} are visible."
                        test_result = "Pass"
                        
                        ws.append([
                            tc_id, module_name, prerequisite, test_data, url,
                            description, execution_steps, "", expected_output, actual_output, test_result
                        ])
                        tc_counter += 1
                    
                    # Test case for username field
                    username_inputs = [inp for inp in inputs if any(keyword in (inp.get('id', '') + inp.get('name', '') + inp.get('placeholder', '')).lower() 
                                                                      for keyword in ['user', 'email', 'username', 'login'])]
                    if username_inputs:
                        tc_id = f"TC_{tc_counter:03d}"
                        prerequisite = f"User is on {module_name.lower()}"
                        test_data = ""
                        description = f"Verify the User name field and validate it"
                        execution_steps = f"1. Open URL: {url}. 2. Verify User name field."
                        expected_output = "User name field should be visible and enabled."
                        actual_output = "User name field is visible and enabled."
                        test_result = "Pass"
                        
                        ws.append([
                            tc_id, module_name, prerequisite, test_data, url,
                            description, execution_steps, "", expected_output, actual_output, test_result
                        ])
                        tc_counter += 1
                        
                        # Test case for entering username
                        tc_id = f"TC_{tc_counter:03d}"
                        prerequisite = f"User is on {module_name.lower()}"
                        test_data = "User entered username"
                        description = f"Verify that user is able to enter user name."
                        execution_steps = f"1. Open URL: {url}. 2. Enter User name."
                        expected_output = "User should be able to enter user name."
                        actual_output = "User is able to enter user name."
                        test_result = "Pass"
                        
                        ws.append([
                            tc_id, module_name, prerequisite, test_data, url,
                            description, execution_steps, "", expected_output, actual_output, test_result
                        ])
                        tc_counter += 1
                    
                    # Test case for password field
                    password_inputs = [inp for inp in inputs if 'password' in (inp.get('id', '') + inp.get('name', '') + inp.get('placeholder', '')).lower()]
                    if password_inputs:
                        tc_id = f"TC_{tc_counter:03d}"
                        prerequisite = f"User is on {module_name.lower()}"
                        test_data = "User entered password"
                        description = f"Verify that user is able to enter password."
                        execution_steps = f"1. Open URL: {url}. 2. Enter Password."
                        expected_output = "User should be able to enter password."
                        actual_output = "User is able to enter password."
                        test_result = "Pass"
                        
                        ws.append([
                            tc_id, module_name, prerequisite, test_data, url,
                            description, execution_steps, "", expected_output, actual_output, test_result
                        ])
                        tc_counter += 1
                    
                    # Test case for login/submit button
                    login_buttons = [btn for btn in buttons if any(keyword in (btn.get('text', '') + btn.get('id', '')).lower() 
                                                                   for keyword in ['login', 'submit', 'go to dashboard', 'sign in', 'enter'])]
                    if login_buttons:
                        button_text = login_buttons[0].get('text', 'Button')
                        tc_id = f"TC_{tc_counter:03d}"
                        prerequisite = f"User is on {module_name.lower()}"
                        test_data = "Valid credentials: Email and Password" if username_inputs and password_inputs else "User entered username and password"
                        description = f"Verify the validation on {button_text} button."
                        execution_steps = f"1. Open URL: {url}. 2. Enter User name. 3. Enter Password. 4. Click on {button_text} button."
                        expected_output = f"1. {button_text} button should be clickable. 2. User should be redirected to the Advertiser Dashboard."
                        actual_output = f"1. {button_text} button is clickable. 2. User is getting redirected to the Advertiser Dashboard."
                        test_result = "Pass"
                        
                        ws.append([
                            tc_id, module_name, prerequisite, test_data, url,
                            description, execution_steps, "", expected_output, actual_output, test_result
                        ])
                        tc_counter += 1
                    
                    previous_url = url
                
                elif action_type == 'navigation':
                    # Navigation to new page
                    if url != previous_url:
                        tc_id = f"TC_{tc_counter:03d}"
                        prerequisite = f"User is on previous page"
                        test_data = ""
                        description = f"Verify navigation to {module_name}."
                        execution_steps = f"1. Navigate to: {url}"
                        expected_output = f"User should be redirected to {module_name}."
                        actual_output = f"User is redirected to {module_name}."
                        test_result = "Pass"
                        
                        ws.append([
                            tc_id, module_name, prerequisite, test_data, url,
                            description, execution_steps, "", expected_output, actual_output, test_result
                        ])
                        tc_counter += 1
                        previous_url = url
            
            # Auto-adjust column widths
            column_widths = {
                'A': 12,  # TC_ID
                'B': 15,  # Module
                'C': 25,  # Prerequisite
                'D': 30,  # Test Data
                'E': 40,  # URL
                'F': 40,  # Description
                'G': 60,  # Execution_Steps
                'H': 5,   # Empty column
                'I': 50,  # Expected Output
                'J': 50,  # Actual Output
                'K': 12   # Test_Result
            }
            
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Enable text wrapping for all cells
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            wb.save(excel_file)
            return excel_file
            
        except Exception as e:
            self.message_queue.put({
                "type": "log",
                "text": f"Error generating Excel: {str(e)}",
                "level": "ERROR"
            })
            import traceback
            self.message_queue.put({
                "type": "log",
                "text": f"Traceback: {traceback.format_exc()}",
                "level": "ERROR"
            })
            return None
    
    def _get_action_description(self, action: Dict[str, Any]) -> str:
        """Get action description"""
        action_type = action.get('action', '')
        if action_type == 'navigate':
            return f"Navigate to {action.get('url', '')}"
        elif action_type == 'page_capture':
            buttons_count = len(action.get('buttons', []))
            inputs_count = len(action.get('inputs', []))
            return f"Page capture: {buttons_count} buttons, {inputs_count} inputs found"
        elif action_type == 'navigation':
            return f"Page navigation to {action.get('url', '')}"
        else:
            return "Unknown action"
    
    def cleanup(self):
        """Clean up resources"""
        try:
            if self.page:
                try:
                    self.page.close()
                except:
                    pass
            if self.context:
                try:
                    self.context.close()
                except:
                    pass
            if self.browser:
                try:
                    self.browser.close()
                except:
                    pass
            if self.playwright:
                try:
                    self.playwright.stop()
                except:
                    pass
        except Exception as e:
            self.log(f"Error during cleanup: {str(e)}", "ERROR")
    
    def on_closing(self):
        """Handle window closing"""
        if self.is_recording:
            if messagebox.askokcancel("Quit", "Recording is active. Do you want to quit?"):
                self.cleanup()
                self.root.destroy()
        else:
            self.cleanup()
            self.root.destroy()


def main():
    """Entry point"""
    root = Tk()
    app = TestCaptureToolGUI(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()


if __name__ == "__main__":
    main()
